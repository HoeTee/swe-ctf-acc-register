#!/usr/bin/env python3
import json
import os
import re
from copy import deepcopy
from dataclasses import dataclass
from pathlib import Path
from typing import Annotated, Any, Callable

import httpx
import typer
import yaml
from openpyxl import Workbook, load_workbook


DEFAULT_CONFIG = Path("config.yaml")
app = typer.Typer(no_args_is_help=True, help="Batch-register GZCTF accounts from registration Excel files.")


@dataclass
class RegistrationRow:
    row: int
    name: str
    phone: str
    unit: str
    department: str
    username: str
    email: str
    password: str


def load_config(path: Path) -> dict[str, Any]:
    config = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    config["_config_path"] = str(path.resolve())
    config["_base_dir"] = str(path.resolve().parent)
    return config


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text[:-2] if text.endswith(".0") else text


def normalize_phone(value: Any) -> str:
    return re.sub(r"\D+", "", normalize_text(value))


def normalize_name(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value))


def render_account_value(pattern: str, phone: str, name: str, email_domain: str) -> str:
    if pattern == "phone":
        return phone
    if pattern == "phone_last4":
        return phone[-4:]
    if pattern == "phone_last6":
        return phone[-6:]
    if pattern in {"phone_email", "email"}:
        return f"{phone}@{email_domain}"
    if pattern == "name":
        return name
    return pattern.format(
        phone=phone,
        phone_last4=phone[-4:],
        phone_last6=phone[-6:],
        name=name,
        email_domain=email_domain,
    )


def render_row_value(pattern: Any, account: RegistrationRow) -> Any:
    if pattern is None:
        return None
    text = str(pattern)
    if text == "name":
        return account.name
    if text == "phone":
        return account.phone
    if text == "username":
        return account.username
    if text == "email":
        return account.email
    return text.format(
        row=account.row,
        name=account.name,
        phone=account.phone,
        unit=account.unit,
        department=account.department,
        username=account.username,
        email=account.email,
    )


def resolve_path(path_value: str, base_dir: Path) -> Path:
    path = Path(path_value)
    if path.is_absolute():
        return path
    return base_dir / path


def find_header(rows: list[tuple[Any, ...]], required: list[str]) -> tuple[int, list[str]]:
    for index, row in enumerate(rows[:30]):
        headers = [normalize_text(value) for value in row]
        if all(any(req == h or req in h for h in headers) for req in required):
            return index, headers
    raise RuntimeError(f"cannot find header row with columns: {required}")


def column_index(headers: list[str], name: str) -> int:
    if name in headers:
        return headers.index(name)
    matches = [index for index, value in enumerate(headers) if name in value]
    if len(matches) == 1:
        return matches[0]
    raise RuntimeError(f"cannot locate column {name!r}; headers={headers!r}")


def merge_registration_config(config: dict[str, Any], cluster: dict[str, Any]) -> dict[str, Any]:
    defaults = dict(config.get("registration_defaults") or {})
    legacy = dict(config.get("registration") or {})
    cluster_registration = dict(cluster.get("registration") or {})
    merged = {**defaults, **legacy, **cluster_registration}
    if not merged.get("xlsx_path"):
        raise RuntimeError(f"cluster {cluster.get('name')} missing registration.xlsx_path")
    return merged


def load_registration(
    config: dict[str, Any],
    cluster: dict[str, Any] | None = None,
) -> tuple[list[RegistrationRow], list[dict[str, Any]], dict[str, Any]]:
    base_dir = Path(config["_base_dir"])
    reg_cfg = merge_registration_config(config, cluster or {})
    account_cfg = config.get("account", {})
    xlsx_path = resolve_path(reg_cfg["xlsx_path"], base_dir)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[reg_cfg.get("sheet_name")] if reg_cfg.get("sheet_name") else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    required = [
        reg_cfg["name_column"],
        reg_cfg["unit_column"],
        reg_cfg["department_column"],
        reg_cfg["phone_column"],
    ]
    header_idx, headers = find_header(rows, required)
    name_idx = column_index(headers, reg_cfg["name_column"])
    unit_idx = column_index(headers, reg_cfg["unit_column"])
    dept_idx = column_index(headers, reg_cfg["department_column"])
    phone_idx = column_index(headers, reg_cfg["phone_column"])

    email_domain = str(account_cfg.get("email_domain", "zjrcu.com")).strip().lstrip("@")
    username_pattern = str(account_cfg.get("username", "phone"))
    email_pattern = str(account_cfg.get("email", "phone_email"))
    password_pattern = str(account_cfg.get("password", "zjrcu{phone_last4}"))
    accounts: list[RegistrationRow] = []
    warnings: list[dict[str, Any]] = []
    seen_usernames: dict[str, int] = {}

    for row_number, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        name = normalize_name(row[name_idx] if name_idx < len(row) else None)
        phone = normalize_phone(row[phone_idx] if phone_idx < len(row) else None)
        unit = normalize_text(row[unit_idx] if unit_idx < len(row) else None)
        department = normalize_text(row[dept_idx] if dept_idx < len(row) else None)
        if not name and not phone:
            continue
        if not name:
            warnings.append({"type": "missing_name", "row": row_number, "phone": phone})
            continue
        if len(phone) < 4:
            warnings.append({"type": "invalid_phone", "row": row_number, "name": name, "phone": phone})
            continue
        username = render_account_value(username_pattern, phone=phone, name=name, email_domain=email_domain)
        email = render_account_value(email_pattern, phone=phone, name=name, email_domain=email_domain)
        password = render_account_value(password_pattern, phone=phone, name=name, email_domain=email_domain)
        if username in seen_usernames:
            warnings.append(
                {
                    "type": "duplicate_username",
                    "row": row_number,
                    "first_row": seen_usernames[username],
                    "username": username,
                }
            )
            continue
        seen_usernames[username] = row_number
        accounts.append(
            RegistrationRow(
                row=row_number,
                name=name,
                phone=phone,
                unit=unit,
                department=department,
                username=username,
                email=email,
                password=password,
            )
        )

    source = {
        "xlsx_path": str(xlsx_path),
        "sheet_name": ws.title,
        "total_accounts": len(accounts),
    }
    return accounts, warnings, source


class GZCTFClient:
    def __init__(self, cluster: dict[str, Any]):
        self.cluster = cluster
        self.name = str(cluster["name"])
        self.base_url = str(cluster["base_url"]).rstrip("/")
        self.client = httpx.Client(base_url=self.base_url, timeout=float(cluster.get("timeout_seconds", 15)))

    def close(self) -> None:
        self.client.close()

    def clear_session(self) -> None:
        self.client.cookies.clear()

    def admin_credentials(self) -> tuple[str, str]:
        username = self.cluster.get("admin_username")
        password = self.cluster.get("admin_password")
        if self.cluster.get("admin_username_env"):
            username = os.environ.get(str(self.cluster["admin_username_env"]))
        if self.cluster.get("admin_password_env"):
            password = os.environ.get(str(self.cluster["admin_password_env"]))
        if not username or not password:
            raise RuntimeError(
                f"missing admin credentials for {self.name}; set admin_username/password or env vars"
            )
        return str(username), str(password)

    def login_admin(self) -> None:
        username, password = self.admin_credentials()
        response = self.client.post("/api/account/login", json={"userName": username, "password": password})
        response.raise_for_status()

    def login_user(self, account: RegistrationRow) -> dict[str, Any]:
        try:
            response = self.client.post(
                "/api/account/login",
                json={"userName": account.username, "password": account.password},
            )
        except httpx.TimeoutException as exc:
            return {"status": "failed", "http_status": None, "message": f"timeout: {exc}"}
        except httpx.HTTPError as exc:
            return {"status": "failed", "http_status": None, "message": f"http_error: {exc}"}
        if response.status_code in {200, 201, 204}:
            return {"status": "logged_in", "http_status": response.status_code}
        return {"status": "failed", "http_status": response.status_code, "message": response.text[:1000]}

    def get_platform_config(self) -> dict[str, Any]:
        response = self.client.get("/api/admin/config")
        response.raise_for_status()
        payload = response.json()
        if not isinstance(payload, dict):
            raise RuntimeError(f"unexpected /api/admin/config response: {payload!r}")
        return payload

    def update_platform_config(self, config_payload: dict[str, Any], method: str = "auto") -> str:
        methods = ["put", "patch", "post"] if method == "auto" else [method.lower()]
        last_error = None
        for candidate in methods:
            request = getattr(self.client, candidate, None)
            if request is None:
                continue
            response = request("/api/admin/config", json=config_payload)
            if response.status_code < 400:
                return candidate.upper()
            last_error = f"{candidate.upper()} /api/admin/config -> HTTP {response.status_code}: {response.text[:500]}"
            if response.status_code not in {404, 405, 415}:
                response.raise_for_status()
        raise RuntimeError(last_error or "no config update method tried")

    def register_account(self, account: RegistrationRow) -> dict[str, Any]:
        payload = {
            "challenge": None,
            "userName": account.username,
            "password": account.password,
            "email": account.email,
        }
        try:
            response = self.client.post("/api/account/register", json=payload)
        except httpx.TimeoutException as exc:
            return {"status": "failed", "http_status": None, "message": f"timeout: {exc}"}
        except httpx.HTTPError as exc:
            return {"status": "failed", "http_status": None, "message": f"http_error: {exc}"}
        text = response.text
        if response.status_code in {200, 201, 204}:
            return {"status": "created", "http_status": response.status_code}
        lowered = text.lower()
        if response.status_code in {400, 409} and any(
            marker in lowered for marker in ["duplicate", "already", "exist", "taken"]
        ):
            return {"status": "already_exists", "http_status": response.status_code, "message": text[:500]}
        return {"status": "failed", "http_status": response.status_code, "message": text[:1000]}

    def create_team(self, account: RegistrationRow, team_config: dict[str, Any]) -> dict[str, Any]:
        name = render_row_value(team_config.get("name", "name"), account)
        bio = render_row_value(team_config.get("bio"), account)
        payload = {"name": name, "bio": bio}
        try:
            response = self.client.post("/api/team", json=payload)
        except httpx.TimeoutException as exc:
            return {
                "team_status": "failed",
                "team_http_status": None,
                "team_name": name,
                "team_message": f"timeout: {exc}",
            }
        except httpx.HTTPError as exc:
            return {
                "team_status": "failed",
                "team_http_status": None,
                "team_name": name,
                "team_message": f"http_error: {exc}",
            }
        text = response.text
        if response.status_code in {200, 201, 204}:
            return {"team_status": "created", "team_http_status": response.status_code, "team_name": name}
        lowered = text.lower()
        if response.status_code in {400, 409} and any(
            marker in lowered
            for marker in ["duplicate", "already", "exist", "taken", "joined", "already in a team"]
        ):
            return {
                "team_status": "already_exists",
                "team_http_status": response.status_code,
                "team_name": name,
                "team_message": text[:500],
            }
        return {
            "team_status": "failed",
            "team_http_status": response.status_code,
            "team_name": name,
            "team_message": text[:1000],
        }


def account_to_public_dict(account: RegistrationRow, include_password: bool = False) -> dict[str, Any]:
    value = {
        "row": account.row,
        "name": account.name,
        "phone": account.phone,
        "unit": account.unit,
        "department": account.department,
        "username": account.username,
        "email": account.email,
    }
    if include_password:
        value["password"] = account.password
    return value


def write_result_workbook(output: dict[str, Any], path: Path) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "summary"
    summary.append(
        [
            "cluster",
            "base_url",
            "registration_file",
            "registration_sheet",
            "total_planned",
            "created",
            "already_exists",
            "failed",
            "teams_created",
            "teams_already_exists",
            "teams_failed",
            "dry_run",
            "config_restored",
        ]
    )
    details = wb.create_sheet("accounts")
    details.append(
        [
            "cluster",
            "row",
            "name",
            "phone",
            "unit",
            "department",
            "username",
            "email",
            "password",
            "status",
            "http_status",
            "message",
            "team_name",
            "team_status",
            "team_http_status",
            "team_message",
        ]
    )
    warnings_sheet = wb.create_sheet("warnings")
    warnings_sheet.append(["cluster", "type", "row", "name", "phone", "message"])

    for cluster_result in output.get("clusters", []):
        source = cluster_result.get("registration_source") or {}
        summary.append(
            [
                cluster_result.get("cluster"),
                cluster_result.get("base_url"),
                source.get("xlsx_path"),
                source.get("sheet_name"),
                cluster_result.get("total_planned"),
                cluster_result.get("created"),
                cluster_result.get("already_exists"),
                cluster_result.get("failed"),
                cluster_result.get("teams_created"),
                cluster_result.get("teams_already_exists"),
                cluster_result.get("teams_failed"),
                cluster_result.get("dry_run"),
                cluster_result.get("config_restored"),
            ]
        )
        for item in cluster_result.get("items", []):
            details.append(
                [
                    cluster_result.get("cluster"),
                    item.get("row"),
                    item.get("name"),
                    item.get("phone"),
                    item.get("unit"),
                    item.get("department"),
                    item.get("username"),
                    item.get("email"),
                    item.get("password"),
                    item.get("status"),
                    item.get("http_status"),
                    item.get("message"),
                    item.get("team_name"),
                    item.get("team_status"),
                    item.get("team_http_status"),
                    item.get("team_message"),
                ]
            )
        for warning in cluster_result.get("warnings", []):
            warnings_sheet.append(
                [
                    cluster_result.get("cluster"),
                    warning.get("type"),
                    warning.get("row"),
                    warning.get("name"),
                    warning.get("phone"),
                    warning.get("message") or json.dumps(warning, ensure_ascii=False),
                ]
            )

    for worksheet in wb.worksheets:
        for column_cells in worksheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 48)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_result_files(config: dict[str, Any], output: dict[str, Any]) -> None:
    base_dir = Path(config["_base_dir"])
    output_config = config.get("output") or {}
    result_file = output_config.get("result_file")
    if result_file:
        path = resolve_path(result_file, base_dir)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(output, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    xlsx_file = output_config.get("xlsx_file")
    if xlsx_file:
        write_result_workbook(output, resolve_path(xlsx_file, base_dir))


def compact_message(value: Any, max_len: int = 240) -> str:
    text = normalize_text(value).replace("\r", " ").replace("\n", " ")
    return text[:max_len] + "..." if len(text) > max_len else text


def build_enabled_registration_config(original: dict[str, Any], updates: dict[str, Any]) -> dict[str, Any]:
    payload = deepcopy(original)
    for key, value in updates.items():
        payload[key] = value
    return payload


def run_cluster(
    cluster: dict[str, Any],
    accounts: list[RegistrationRow],
    registration_settings: dict[str, Any],
    dry_run: bool,
    limit: int | None,
    keep_register_open: bool,
    progress_every: int,
    failure_details: int,
    team_config: dict[str, Any],
    on_update: Callable[[dict[str, Any]], None] | None = None,
) -> dict[str, Any]:
    selected_accounts = accounts[:limit] if limit else accounts
    result = {
        "cluster": cluster["name"],
        "base_url": cluster["base_url"],
        "registration_source": cluster.get("_registration_source", {}),
        "warnings": cluster.get("_registration_warnings", []),
        "dry_run": dry_run,
        "total_planned": len(selected_accounts),
        "created": 0,
        "already_exists": 0,
        "failed": 0,
        "teams_created": 0,
        "teams_already_exists": 0,
        "teams_failed": 0,
        "items": [],
        "config_update_method": None,
        "config_restored": False,
    }

    if dry_run:
        result["items"] = [
            {
                **account_to_public_dict(account, include_password=True),
                "status": "planned",
                "team_name": render_row_value(team_config.get("name", "name"), account),
                "team_status": "planned" if team_config.get("create", True) else "skipped",
            }
            for account in selected_accounts
        ]
        if on_update:
            on_update(result)
        return result

    client = GZCTFClient(cluster)
    original_config = None
    shown_failures = 0
    try:
        typer.echo(f"[{cluster['name']}] login admin")
        client.login_admin()
        typer.echo(f"[{cluster['name']}] read platform config")
        original_config = client.get_platform_config()
        enabled_config = build_enabled_registration_config(original_config, registration_settings)
        typer.echo(f"[{cluster['name']}] enable registration")
        result["config_update_method"] = client.update_platform_config(
            enabled_config,
            method=str(cluster.get("config_update_method", "auto")),
        )
        for index, account in enumerate(selected_accounts, start=1):
            client.clear_session()
            register_result = client.register_account(account)
            status = register_result["status"]
            item = {
                **account_to_public_dict(account, include_password=True),
                **register_result,
            }
            if status == "created":
                result["created"] += 1
            elif status == "already_exists":
                result["already_exists"] += 1
            else:
                result["failed"] += 1

            if team_config.get("create", True) and status in {"created", "already_exists"}:
                login_result = client.login_user(account)
                if login_result["status"] == "logged_in":
                    team_result = client.create_team(account, team_config)
                else:
                    team_result = {
                        "team_status": "failed",
                        "team_http_status": login_result.get("http_status"),
                        "team_name": render_row_value(team_config.get("name", "name"), account),
                        "team_message": f"user login failed: {login_result.get('message')}",
                    }
                item.update(team_result)
                if item["team_status"] == "created":
                    result["teams_created"] += 1
                elif item["team_status"] == "already_exists":
                    result["teams_already_exists"] += 1
                else:
                    result["teams_failed"] += 1
            elif not team_config.get("create", True):
                item["team_status"] = "skipped"

            result["items"].append(item)
            if status == "failed" and shown_failures < failure_details:
                shown_failures += 1
                typer.echo(
                    f"[{cluster['name']}] failed row={account.row} "
                    f"name={account.name} username={account.username} "
                    f"http={register_result.get('http_status')} "
                    f"message={compact_message(register_result.get('message'))}",
                    err=True,
                )
            if item.get("team_status") == "failed" and shown_failures < failure_details:
                shown_failures += 1
                typer.echo(
                    f"[{cluster['name']}] team failed row={account.row} "
                    f"name={account.name} username={account.username} "
                    f"http={item.get('team_http_status')} "
                    f"message={compact_message(item.get('team_message'))}",
                    err=True,
                )
            if on_update:
                on_update(result)
            if progress_every > 0 and (index == 1 or index % progress_every == 0 or index == len(selected_accounts)):
                typer.echo(
                    f"[{cluster['name']}] {index}/{len(selected_accounts)} "
                    f"created={result['created']} already_exists={result['already_exists']} failed={result['failed']} "
                    f"teams_created={result['teams_created']} teams_already_exists={result['teams_already_exists']} "
                    f"teams_failed={result['teams_failed']}"
                )
    except Exception as exc:
        result["error"] = str(exc)
        if on_update:
            on_update(result)
        raise
    finally:
        if original_config is not None and not keep_register_open:
            try:
                typer.echo(f"[{cluster['name']}] restore platform config")
                client.login_admin()
                client.update_platform_config(
                    original_config,
                    method=str(cluster.get("config_update_method", "auto")),
                )
                result["config_restored"] = True
                result.pop("config_restore_error", None)
            except Exception as exc:
                result["config_restore_error"] = str(exc)
        client.close()
        if on_update:
            on_update(result)

    return result


def echo_json(value: dict[str, Any]) -> None:
    typer.echo(json.dumps(value, ensure_ascii=False, indent=2))


@app.command("check-excel")
def check_excel(
    config_path: Annotated[Path, typer.Option("--config", "-c", help="Path to config.yaml.")] = DEFAULT_CONFIG,
    print_passwords: Annotated[bool, typer.Option(help="Print generated passwords in the sample output.")] = False,
) -> None:
    """Read each cluster registration sheet and print generated account samples."""
    config = load_config(config_path)
    output = {
        "warnings": [],
    }
    clusters = config.get("gzctf", {}).get("clusters") or [{"name": "default"}]
    output["clusters"] = []
    for cluster in clusters:
        accounts, warnings, source = load_registration(config, cluster)
        output["clusters"].append(
            {
                "cluster": cluster.get("name", "default"),
                "registration_source": source,
                "total_accounts": len(accounts),
                "warnings": warnings,
                "sample": [
                    account_to_public_dict(account, include_password=print_passwords)
                    for account in accounts[:20]
                ],
            }
        )
    echo_json(output)
    has_warnings = any(cluster.get("warnings") for cluster in output["clusters"])
    if has_warnings:
        raise typer.Exit(1)


def execute_registration(
    config_path: Path,
    dry_run: bool,
    limit: int | None,
    keep_register_open: bool,
    timeout_seconds: float | None,
    progress_every: int,
    failure_details: int,
) -> dict[str, Any]:
    config = load_config(config_path)
    output = {"warnings": []}
    registration_settings = config.get("registration_settings") or {}
    team_config = {"create": True, "name": "name", "bio": None, **(config.get("team") or {})}
    clusters = config.get("gzctf", {}).get("clusters") or []
    if not clusters:
        raise typer.BadParameter("config.gzctf.clusters is empty")

    output["clusters"] = []
    for cluster in clusters:
        accounts, warnings, source = load_registration(config, cluster)
        cluster = dict(cluster)
        if timeout_seconds is not None:
            cluster["timeout_seconds"] = timeout_seconds
        cluster["_registration_warnings"] = warnings
        cluster["_registration_source"] = source
        def persist_current(current_result: dict[str, Any]) -> None:
            write_result_files(config, {**output, "clusters": [*output["clusters"], current_result]})

        output["clusters"].append(
            run_cluster(
                cluster=cluster,
                accounts=accounts,
                registration_settings=registration_settings,
                dry_run=dry_run,
                limit=limit,
                keep_register_open=keep_register_open,
                progress_every=progress_every,
                failure_details=failure_details,
                team_config=team_config,
                on_update=persist_current,
            )
        )

    write_result_files(config, output)
    return output


@app.command("dry-run")
def dry_run(
    config_path: Annotated[Path, typer.Option("--config", "-c", help="Path to config.yaml.")] = DEFAULT_CONFIG,
    limit: Annotated[int | None, typer.Option(help="Only process the first N accounts per cluster.")] = None,
) -> None:
    """Generate planned account records without calling GZCTF APIs."""
    echo_json(
        execute_registration(
            config_path,
            dry_run=True,
            limit=limit,
            keep_register_open=False,
            timeout_seconds=None,
            progress_every=50,
            failure_details=0,
        )
    )


@app.command("register")
def register(
    config_path: Annotated[Path, typer.Option("--config", "-c", help="Path to config.yaml.")] = DEFAULT_CONFIG,
    limit: Annotated[int | None, typer.Option(help="Only process the first N accounts per cluster.")] = None,
    keep_register_open: Annotated[
        bool,
        typer.Option(help="Do not restore the original platform registration config."),
    ] = False,
    timeout_seconds: Annotated[
        float | None,
        typer.Option(help="HTTP timeout per request. Overrides cluster.timeout_seconds."),
    ] = None,
    progress_every: Annotated[
        int,
        typer.Option(help="Print progress every N accounts."),
    ] = 10,
    failure_details: Annotated[
        int,
        typer.Option(help="Print details for the first N failed accounts."),
    ] = 5,
) -> None:
    """Register accounts on every configured GZCTF cluster."""
    echo_json(
        execute_registration(
            config_path,
            dry_run=False,
            limit=limit,
            keep_register_open=keep_register_open,
            timeout_seconds=timeout_seconds,
            progress_every=progress_every,
            failure_details=failure_details,
        )
    )


if __name__ == "__main__":
    try:
        app()
    except Exception as exc:
        typer.echo(f"ERROR: {exc}", err=True)
        raise
